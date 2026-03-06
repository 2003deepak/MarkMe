#report service

from fastapi import Request
from fastapi.responses import JSONResponse
from io import BytesIO
from datetime import datetime
import json
import aio_pika

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape, A4

from app.schemas.student import Student
from app.schemas.subject import Subject
from app.schemas.session import Session
from app.schemas.attendance import Attendance
from app.utils.imagekit_uploader import upload_file_to_imagekit
from app.core.rabbitmq_config import settings as rabbit_settings
from app.utils.publisher import send_to_queue


async def download_class_report(
    request: Request,
    department: str,
    program: str,
    semester: int,
    batch_year: int,
    file_type: str = "excel"
):

    #auth
    user = getattr(request.state, "user", None)
    if not user or user.get("role") != "admin":
        return JSONResponse(
            status_code=403,
            content={"message": "Only Admin can access"}
        )

    #students
    students = await Student.find({
        "department": department,
        "program": program,
        "semester": semester,
        "batch_year": batch_year
    }).sort("roll_number").to_list()

    if not students:
        return JSONResponse(status_code=404, content={"message": "No students"})

    student_count = len(students)
    student_rolls = [s.roll_number for s in students]
    student_names = [f"{s.first_name} {s.last_name}" for s in students]

    #subjects
    subjects = await Subject.find({
        "department": department,
        "program": program,
        "semester": semester
    }).to_list()

    subject_map = {str(s.id): s for s in subjects}

    #sessions
    sessions = await Session.find({
        "department": department,
        "program": program,
        "semester": str(semester)
    }, fetch_links=True).to_list()

    if not sessions:
        return JSONResponse(status_code=404, content={"message": "No sessions"})

    session_ids = [s.id for s in sessions]

    #group sessions by subject
    subject_sessions_map = {}
    for s in sessions:
        sid = str(s.subject.id)
        subject_sessions_map.setdefault(sid, []).append(s.id)

    #attendance aggregation
    pipeline = [
        {"$match": {"session.$id": {"$in": session_ids}}},
        {"$project": {"session_id": "$session.$id", "students": 1, "date": 1}}
    ]

    attendances = await Attendance.aggregate(pipeline).to_list()

    attendance_map = {}
    for att in attendances:
        sid = str(att["session_id"])
        attendance_map.setdefault(sid, []).append(att)

    stream = BytesIO()

    # ===================== EXCEL =========================

    if file_type == "excel":

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill("solid", fgColor="366092")
        header_font = Font(color="FFFFFF", bold=True)

        present_fill = PatternFill("solid", fgColor="C6EFCE")
        present_font = Font(color="006100")

        absent_fill = PatternFill("solid", fgColor="FFC7CE")
        absent_font = Font(color="9C0006")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center = Alignment(horizontal="center", vertical="center")

        for subject_id, subject in subject_map.items():

            full_name = f"{subject.subject_name} {subject.component}"

            if len(full_name) > 31:
                words = subject.subject_name.split()
                acronym = "".join(word[0].upper() for word in words if word)
                sheet_name = f"{acronym} {subject.component}"
            else:
                sheet_name = full_name

            ws = wb.create_sheet(title=sheet_name[:31])

            session_ids_for_sub = subject_sessions_map.get(subject_id, [])
            student_data = [dict() for _ in range(student_count)]
            all_dates = set()

            for sid in session_ids_for_sub:
                for att in attendance_map.get(str(sid), []):

                    if not att["students"]:
                        continue

                    date = att["date"].date()
                    bitmask = att["students"].strip()
                    all_dates.add(date)

                    for i in range(min(len(bitmask), student_count)):
                        status = "P" if bitmask[i] == "1" else "A"
                        student_data[i][date] = status

            sorted_dates = sorted(all_dates)

            header = ["Roll No", "Name"] + \
                     [d.strftime("%d-%m-%Y") for d in sorted_dates] + \
                     ["Total", "Present", "%"]

            ws.append(header)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            for i in range(student_count):

                total = len(sorted_dates)
                present_count = 0
                row = [student_rolls[i], student_names[i]]

                for d in sorted_dates:
                    val = student_data[i].get(d, "A")
                    if val == "P":
                        present_count += 1
                    row.append(val)

                percent = (present_count / total * 100) if total else 0
                row.extend([total, present_count, f"{percent:.1f}%"])
                ws.append(row)

                r = i + 2

                for c in range(1, len(row) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border

                    if 3 <= c <= 2 + total:
                        if cell.value == "P":
                            cell.fill = present_fill
                            cell.font = present_font
                        else:
                            cell.fill = absent_fill
                            cell.font = absent_font
                        cell.alignment = center
                    else:
                        cell.alignment = center

            ws.freeze_panes = "C2"

            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 3

        wb.save(stream)
        extension = "xlsx"

    # ===================== PDF ==========================

    elif file_type == "pdf":

        doc = SimpleDocTemplate(stream, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        for subject_id, subject in subject_map.items():

            elements.append(Paragraph(
                f"{subject.subject_name} - {subject.component}",
                styles["Heading2"]
            ))
            elements.append(Spacer(1, 12))

            session_ids_for_sub = subject_sessions_map.get(subject_id, [])
            student_data = [dict() for _ in range(student_count)]
            all_dates = set()

            for sid in session_ids_for_sub:
                for att in attendance_map.get(str(sid), []):

                    if not att["students"]:
                        continue

                    date = att["date"].date()
                    bitmask = att["students"].strip()
                    all_dates.add(date)

                    for i in range(min(len(bitmask), student_count)):
                        status = "P" if bitmask[i] == "1" else "A"
                        student_data[i][date] = status

            sorted_dates = sorted(all_dates)

            table_data = [["Roll No", "Name"] +
                          [d.strftime("%d-%m-%Y") for d in sorted_dates] +
                          ["Total", "Present", "%"]]

            for i in range(student_count):

                total = len(sorted_dates)
                present_count = 0
                row = [student_rolls[i], student_names[i]]

                for d in sorted_dates:
                    val = student_data[i].get(d, "A")
                    if val == "P":
                        present_count += 1
                    row.append(val)

                percent = (present_count / total * 100) if total else 0
                row.extend([total, present_count, f"{percent:.1f}%"])

                table_data.append(row)

            table = Table(table_data, repeatRows=1)

            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (2, 1), (-1, -1), "CENTER")
            ]))

            elements.append(table)
            elements.append(Spacer(1, 24))

        doc.build(elements)
        extension = "pdf"

    else:
        return JSONResponse(status_code=400, content={"message": "Invalid file type"})

    # =================== UPLOAD ==========================

    stream.seek(0)
    filename = f"ERP_Report_{int(datetime.now().timestamp())}.{extension}"

    upload_result = await upload_file_to_imagekit(
        file=stream.getvalue(),
        filename=filename,
        folder="reports",
        tags=["erp", "attendance"]
    )
    
    # ============= PUSH DELAYED DELETE JOB ==============
    
    # 24 hours in milliseconds
    delay_ms = 24 * 60 * 60 * 1000  

    await send_to_queue(
        queue_name="cleanup_queue",
        payload={
            "type": "delete_file",
            "data": {
                "file_id": upload_result["fileId"]
            }
        },
        priority=10,
        delay_ms=delay_ms
    )


    # ================== RESPONSE =========================

    return JSONResponse(
        status_code=200,
        content={
            "success": True,
            "message": "Report generated successfully",
            "file_url": upload_result["url"],
            "file_id": upload_result["fileId"],
        }
    )